#!/usr/bin/env python3
"""
Excel Bakım Planı Parser
Excel'deki bakım planını JSON formatına çevirir
"""

import openpyxl
import json
from collections import defaultdict
from pathlib import Path

# Excel dosya yolu
EXCEL_PATH = Path(r"C:\Users\Onars\OneDrive\Masaüstü\Siveno Geliştirmeler\EK-3 YILLIK BAKIM PLANI 2024.xlsx")
OUTPUT_JSON = Path(__file__).parent / "machines_data.json"

# Ay yapısı: her ay 4 hücre (4 hafta)
MONTH_STRUCTURE = {
    1: (4, 7),    # Ocak
    2: (8, 11),   # Şubat
    3: (12, 15),  # Mart
    4: (16, 19),  # Nisan
    5: (20, 23),  # Mayıs
    6: (24, 27),  # Haziran
    7: (28, 31),  # Temmuz
    8: (32, 35),  # Ağustos
    9: (36, 39),  # Eylül
    10: (40, 43), # Ekim
    11: (44, 47), # Kasım
    12: (48, 51)  # Aralık
}

MONTH_NAMES = [
    'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
    'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık'
]

# Renk kodları -> Periyot eşleştirmesi
COLOR_TO_PERIOD = {
    'FFFF0000': {'name': 'Yıllık', 'frequency': 'annual'},
    'FFC00000': {'name': '6 Aylık', 'frequency': 'semi-annual'},
    'THEME_3': {'name': '6 Aylık', 'frequency': 'semi-annual'},
    'FF00B050': {'name': '3 Aylık', 'frequency': 'quarterly'},
    'FFFFFF00': {'name': 'Aylık', 'frequency': 'monthly'},
    'THEME_9': {'name': 'Haftalık', 'frequency': 'weekly'},
}


def get_color_code(cell):
    """Hücrenin renk kodunu string olarak döndür"""
    if not cell.fill or not cell.fill.start_color:
        return None

    if cell.fill.patternType != 'solid':
        return None

    color = cell.fill.start_color

    # RGB string
    if hasattr(color, 'rgb') and isinstance(color.rgb, str):
        return color.rgb

    # Theme renkleri
    if hasattr(color, 'type') and color.type == 'theme' and hasattr(color, 'theme'):
        return f'THEME_{color.theme}'

    return None


def parse_excel():
    """Excel dosyasını parse et ve JSON formatına çevir"""

    print(f"Excel dosyasi okunuyor: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    all_machines = []
    machine_map = {}  # machine_no -> machine dict (tekil tutmak için)

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Sayfa1':
            continue

        print(f"\nSayfa: {sheet_name}")
        ws = wb[sheet_name]

        for row_idx in range(8, 60):
            machine_no_cell = ws.cell(row=row_idx, column=1).value
            if not machine_no_cell or str(machine_no_cell).strip() == '':
                continue

            machine_no = str(machine_no_cell).strip()
            machine_name = str(ws.cell(row=row_idx, column=2).value or '').strip()
            bakim_type = str(ws.cell(row=row_idx, column=3).value or '').strip()

            # Her periyot tipi için ayları topla
            schedules = defaultdict(list)

            for month_num, (start_col, end_col) in MONTH_STRUCTURE.items():
                # İlk haftanın rengine bak
                cell = ws.cell(row=row_idx, column=start_col)
                color_code = get_color_code(cell)

                ignore_colors = ['00000000', '00FFFFFF', 'FFFFFFFF', 'FF000000']
                if color_code and color_code not in ignore_colors:
                    if color_code in COLOR_TO_PERIOD:
                        period_info = COLOR_TO_PERIOD[color_code]
                        schedules[period_info['frequency']].append(month_num)

            if schedules:
                # Makineyi ekle veya güncelle
                if machine_no not in machine_map:
                    machine_map[machine_no] = {
                        'machine_no': machine_no,
                        'machine_name': machine_name,
                        'location': None,  # Excel'de lokasyon bilgisi yok
                        'is_active': True,
                        'schedules': []
                    }

                # Schedule'ları ekle
                for frequency, months in schedules.items():
                    machine_map[machine_no]['schedules'].append({
                        'maintenance_type': bakim_type,
                        'frequency': frequency,
                        'months': sorted(months),
                        'is_active': True
                    })

    # Machine map'i listeye çevir
    all_machines = list(machine_map.values())

    # İstatistikler
    total_machines = len(all_machines)
    total_schedules = sum(len(m['schedules']) for m in all_machines)

    print(f"\nParse tamamlandi!")
    print(f"   Toplam makine: {total_machines}")
    print(f"   Toplam schedule: {total_schedules}")

    # Periyot bazlı istatistik
    by_frequency = defaultdict(int)
    for m in all_machines:
        for s in m['schedules']:
            by_frequency[s['frequency']] += 1

    print(f"\nPeriyot Dagilimi:")
    freq_names = {
        'monthly': 'Aylık',
        'quarterly': '3 Aylık',
        'semi-annual': '6 Aylık',
        'annual': 'Yıllık',
        'weekly': 'Haftalık'
    }
    for freq in ['monthly', 'quarterly', 'semi-annual', 'annual', 'weekly']:
        if freq in by_frequency:
            print(f"   {freq_names[freq]:10s}: {by_frequency[freq]:3d} schedule")

    return {
        'machines': all_machines,
        'stats': {
            'total_machines': total_machines,
            'total_schedules': total_schedules,
            'by_frequency': dict(by_frequency)
        }
    }


def save_json(data):
    """JSON dosyasına kaydet"""
    print(f"\nJSON dosyasi olusturuluyor: {OUTPUT_JSON}")

    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"JSON dosyasi kaydedildi!")


def main():
    """Ana fonksiyon"""
    print("="*80)
    print("EXCEL BAKIM PLANI PARSER")
    print("="*80)

    try:
        data = parse_excel()
        save_json(data)

        print(f"\n{'='*80}")
        print("Islem tamamlandi!")
        print(f"{'='*80}\n")

        # Örnek makine göster
        if data['machines']:
            print("Ornek makine:")
            machine = data['machines'][0]
            print(f"   No: {machine['machine_no']}")
            print(f"   Ad: {machine['machine_name']}")
            print(f"   Schedule sayısı: {len(machine['schedules'])}")

            if machine['schedules']:
                schedule = machine['schedules'][0]
                month_names = [MONTH_NAMES[m-1] for m in schedule['months']]
                print(f"\n   İlk Schedule:")
                print(f"      Bakım: {schedule['maintenance_type']}")
                print(f"      Periyot: {schedule['frequency']}")
                print(f"      Aylar: {', '.join(month_names)}")

    except Exception as e:
        print(f"\nHATA: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0


if __name__ == '__main__':
    exit(main())
